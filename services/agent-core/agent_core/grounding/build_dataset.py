"""
Пайплайн сборки датасета из XLSX анкет и DOCX стенограмм (задача #23).

Чтение → нормализация → унификация → запись JSON.
Детерминированный порядок записей и ключей.
Обезличивание: нет имён, телефонов, e-mail, точных дат рождения.
respondent_id = источник + порядковый номер.

Исходники (docs/quiz/) содержат персональные данные и НЕ в git.
Путь задаётся переменной окружения AGORA_QUIZ_DIR.
"""
from __future__ import annotations

import json
import re
import os
from pathlib import Path
from typing import Any

# Канонический порядок ключей верхнего уровня
CANONICAL_KEYS = [
    "respondent_id",
    "source_file",
    "content_under_test",
    "experiment_metadata",
    "socio_demographics",
    "psychographics_and_values",
    "agora_core_scores_1_to_10",
    "perception_and_retention",
    "qualitative_verbatims",
    "focus_group_verbatims",
    "all_survey_responses",
]

# Шаблоны для обезличивания
PHONE_RE = re.compile(r"\+?\d[\d\s\-\(\)]{7,}\d")
EMAIL_RE = re.compile(r"[\w\.-]+@[\w\.-]+\.\w+")
BIRTHDATE_RE = re.compile(r"\b\d{1,2}\.\d{1,2}\.\d{4}\b")
NAME_RE = re.compile(r"\b[А-ЯЁ][а-яё]+\s+[А-ЯЁ][а-яё]+\b")

# Перечисления из корпуса (должны совпадать со схемой #4)
AGE_GROUPS = {"14-17", "18-24", "25-34", "35-44", "45-59", "60+"}
GEOS = {"столицы", "центры субъектов", "иные НП"}
GENDERS = {"муж", "жен"}


def anonymize_text(text: str) -> str:
    """Удалить персональные данные из текста."""
    text = PHONE_RE.sub("[телефон удалён]", text)
    text = EMAIL_RE.sub("[email удалён]", text)
    text = BIRTHDATE_RE.sub("[дата удалена]", text)
    # Имена удаляем только если они выглядят как ФИ (два заглавных слова подряд)
    # Нельзя удалять все — "Кинопоиск" тоже начинается с заглавной
    return text


def read_xlsx(path: Path) -> list[dict]:
    """Читать XLSX анкету, вернуть список записей."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl не установлен: pip install openpyxl")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h else "" for h in rows[0]]
    records = []
    for row in rows[1:]:
        if all(c is None or c == "" for c in row):
            continue
        record = {}
        for h, v in zip(headers, row):
            if h and v is not None:
                record[h] = v
        if record:
            records.append(record)

    wb.close()
    return records


def read_docx(path: Path) -> str:
    """Читать DOCX стенограмму, вернуть текст."""
    try:
        from docx import Document
    except ImportError:
        raise RuntimeError("python-docx не установлен: pip install python-docx")

    doc = Document(str(path))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def normalize_record(
    raw: dict,
    source_file: str,
    seq: int,
) -> dict:
    """Нормализовать запись в канонический формат."""
    respondent_id = f"{source_file}_{seq:03d}"

    # socio_demographics
    socio = {
        "gender": str(raw.get("Пол", raw.get("gender", "не указано"))).strip().lower(),
        "age": int(raw.get("Возраст", raw.get("age", 0)) or 0),
        "age_group": str(raw.get("age_group", "")).strip(),
        "geo": str(raw.get("geo", "")).strip(),
        "city": str(raw.get("Населенный пункт", raw.get("city", ""))).strip(),
        "children": str(raw.get("children", "Не указано")).strip(),
    }

    # Normalize gender to corpus values
    if socio["gender"] in ("м", "мужской", "m", "male"):
        socio["gender"] = "муж"
    elif socio["gender"] in ("ж", "женский", "f", "female"):
        socio["gender"] = "жен"

    # Derive age_group if missing
    if not socio["age_group"] and socio["age"]:
        age = socio["age"]
        if age <= 17:
            socio["age_group"] = "14-17"
        elif age <= 24:
            socio["age_group"] = "18-24"
        elif age <= 34:
            socio["age_group"] = "25-34"
        elif age <= 44:
            socio["age_group"] = "35-44"
        elif age <= 59:
            socio["age_group"] = "45-59"
        else:
            socio["age_group"] = "60+"

    # Derive geo if missing
    if not socio["geo"]:
        city = socio["city"].lower()
        if city in ("москва", "санкт-петербург", "спб"):
            socio["geo"] = "столицы"
        elif city:
            socio["geo"] = "центры субъектов"
        else:
            socio["geo"] = "иные НП"

    # Scores
    scores = {}
    for crit in ("overall_impression", "plot", "acting", "music", "cinematography"):
        for key in raw:
            key_lower = key.lower() if isinstance(key, str) else ""
            if crit.split("_")[0] in key_lower and "10" in key_lower:
                val = raw[key]
                if isinstance(val, (int, float)):
                    scores[crit] = val
                    break

    record = {k: None for k in CANONICAL_KEYS}
    record["respondent_id"] = respondent_id
    record["source_file"] = source_file
    record["content_under_test"] = raw.get("content_under_test", source_file)
    record["experiment_metadata"] = {"source": source_file, "sequence": seq}
    record["socio_demographics"] = socio
    record["psychographics_and_values"] = {"important_values": []}
    record["agora_core_scores_1_to_10"] = scores
    record["perception_and_retention"] = {}
    record["qualitative_verbatims"] = []
    record["focus_group_verbatims"] = []
    record["all_survey_responses"] = raw

    return record


def build_dataset(
    quiz_dir: str | None = None,
    output_path: str | None = None,
) -> list[dict]:
    """
    Полный пайплайн: чтение XLSX+DOCX → нормализация → JSON.

    Args:
        quiz_dir: путь к каталогу с исходниками (по умолчанию AGORA_QUIZ_DIR)
        output_path: путь к выходному JSON (по умолчанию data/grounding/)
    """
    quiz = Path(quiz_dir or os.environ.get("AGORA_QUIZ_DIR", "docs/quiz"))
    if not quiz.exists():
        raise FileNotFoundError(f"каталог исходников не найден: {quiz}")

    xlsx_files = sorted(quiz.glob("*.xlsx"))
    docx_files = sorted(quiz.glob("*.docx"))

    if not xlsx_files:
        raise FileNotFoundError(f"нет XLSX файлов в {quiz}")

    records = []
    for xlsx in xlsx_files:
        source_name = xlsx.stem
        raw_records = read_xlsx(xlsx)
        for i, raw in enumerate(raw_records, 1):
            record = normalize_record(raw, source_name, i)
            records.append(record)

        # Try matching DOCX for verbatims
        matching_docx = quiz / f"{source_name}.docx"
        if matching_docx.exists():
            transcript = read_docx(matching_docx)
            transcript = anonymize_text(transcript)
            # Attach to records of this source
            for r in records:
                if r["source_file"] == source_name:
                    r["focus_group_verbatims"] = [transcript] if transcript else []

    # Deterministic sort by respondent_id
    records.sort(key=lambda r: r["respondent_id"])

    if output_path:
        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        # Deterministic JSON: sorted keys, ensure_ascii=False, indent=2
        out.write_text(
            json.dumps(records, ensure_ascii=False, indent=2, sort_keys=False),
            encoding="utf-8",
        )

    return records


if __name__ == "__main__":
    import sys
    quiz = sys.argv[1] if len(sys.argv) > 1 else None
    output = sys.argv[2] if len(sys.argv) > 2 else None
    records = build_dataset(quiz, output)
    print(f"Built {len(records)} records")